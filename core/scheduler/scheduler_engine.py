import io
from datetime import datetime, timedelta
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

class ContentScheduler:
    @staticmethod
    def distribute_schedule(
        contents: List[Dict[str, Any]],
        start_date: datetime,
        posts_per_day: int,
        time_slots: Dict[str, List[str]]
    ) -> List[Dict[str, Any]]:
        if not contents:
            return []

        # 1. Gom nhóm theo Content Pillar để xếp xen kẽ (Round-robin)
        pillar_groups: Dict[str, List[Dict[str, Any]]] = {}
        for item in contents:
            p_name = item["data"].dna.pillar_name
            pillar_groups.setdefault(p_name, []).append(item)

        balanced_list: List[Dict[str, Any]] = []
        last_pillar = None
        has_items = True

        while has_items:
            has_items = False
            for p_name, group in list(pillar_groups.items()):
                if group and p_name != last_pillar:
                    balanced_list.append(group.pop(0))
                    last_pillar = p_name
                    has_items = True
                    break
            if not has_items:
                for group in pillar_groups.values():
                    if group:
                        balanced_list.append(group.pop(0))
                        has_items = True
                        break

        # 2. Gán ngày và khung giờ cho từng nền tảng
        scheduled_contents = []
        current_date = start_date
        content_idx = 0

        while content_idx < len(balanced_list):
            for i in range(posts_per_day):
                if content_idx >= len(balanced_list):
                    break

                item = balanced_list[content_idx]
                schedule_data = {}

                for platform in ["FACEBOOK", "TIKTOK", "YOUTUBE"]:
                    slots = time_slots.get(platform, ["19:30"])
                    slot_str = slots[i % len(slots)]
                    schedule_data[f"{platform.lower()}_date"] = current_date.strftime("%Y-%m-%d")
                    schedule_data[f"{platform.lower()}_time"] = slot_str

                item["schedule"] = schedule_data
                item["approval_status"] = "APPROVED" if item["data"].score.overall_score >= 8.5 else "PENDING"
                scheduled_contents.append(item)
                content_idx += 1

            current_date += timedelta(days=1)

        return scheduled_contents


class ExcelReportEngine:
    @staticmethod
    def generate_excel(scheduled_data: List[Dict[str, Any]]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Content Matrix Master"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "STT", "Content ID", "File Name", "Google Drive URL", "Content Topic", "Content Pillar",
            "FB Title/Hook", "FB Caption", "FB Hashtags", "FB CTA", "FB Date", "FB Time", "FB Status",
            "TikTok Hook", "TikTok Caption", "TikTok Hashtags", "TikTok CTA", "TikTok Date", "TikTok Time", "TikTok Status",
            "YouTube SEO Title", "YouTube Description", "YouTube Keywords", "YouTube Tags", "YouTube CTA", "YouTube Date", "YouTube Time", "YouTube Status",
            "Overall Status", "Content Score", "Approval", "Notes", "Error Log"
        ]

        ws.append(headers)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for idx, item in enumerate(scheduled_data, 1):
            data = item["data"]
            sched = item.get("schedule", {})

            row_values = [
                idx,
                item["content_id"],
                item["media_name"],
                item["drive_url"],
                item["raw_title"],
                data.dna.pillar_name,
                data.facebook.title_hook,
                data.facebook.caption,
                " ".join(data.facebook.hashtags),
                data.facebook.cta,
                sched.get("facebook_date", ""),
                sched.get("facebook_time", ""),
                "SCHEDULED",
                data.tiktok.hook_caption,
                data.tiktok.hook_caption,
                " ".join(data.tiktok.hashtags),
                data.tiktok.cta,
                sched.get("tiktok_date", ""),
                sched.get("tiktok_time", ""),
                "SCHEDULED",
                data.youtube.seo_title,
                data.youtube.description,
                ", ".join(data.youtube.keywords),
                ", ".join(data.youtube.tags),
                data.youtube.cta,
                sched.get("youtube_date", ""),
                sched.get("youtube_time", ""),
                "SCHEDULED",
                "APPROVED" if item.get("approval_status") == "APPROVED" else "NEEDS_REVIEW",
                data.score.overall_score,
                item.get("approval_status", "PENDING"),
                "",
                ""
            ]
            ws.append(row_values)

            # Format cell & Hyperlink
            current_row = idx + 1
            drive_cell = ws.cell(row=current_row, column=4)
            drive_cell.hyperlink = item["drive_url"]
            drive_cell.font = Font(color="2563EB", underline="single")

            for col in range(1, len(headers) + 1):
                c = ws.cell(row=current_row, column=col)
                c.border = thin_border
                c.alignment = Alignment(vertical="center")

        # Cột Approval Dropdown
        dv = DataValidation(type="list", formula1='"APPROVED,PENDING,REJECTED"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"AE2:AE{len(scheduled_data) + 1}")

        # Tự động điều chỉnh độ rộng cột
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
